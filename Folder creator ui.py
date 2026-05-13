# -*- coding: utf-8 -*-
"""
BIM Folder Creator — single file, import/export CSV & Excel
Run:  python bim_folder_creator.py
Requires for Excel:  pip install openpyxl

Prepared by: Ahmed Khalaf Senior BIM Manager BUJV
"""

import os, csv, tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    XLSX_OK = True
except ImportError:
    XLSX_OK = False

# ── Default data ──────────────────────────────────────────────────────────────
DEFAULT_ROWS = [
    ("1.1.Models Exchange Formats", "1.1.1.RVT"),
    ("1.1.Models Exchange Formats", "1.1.2.NWC"),
    ("1.1.Models Exchange Formats", "1.1.3.IFC"),
    ("1.2.Federated Model",         "1.2.1.With Clash Tests"),
    ("1.2.Federated Model",         "1.2.2.Without Clash Tests"),
    ("1.3.REPORTS & FORMS",         "1.3.1.Clash report"),
    ("1.3.REPORTS & FORMS",         "1.3.2.Model health reports"),
    ("1.3.REPORTS & FORMS",         "1.3.3.Checklists"),
    ("1.3.REPORTS & FORMS",         "1.3.4.Transmittal form"),
    ("1.3.REPORTS & FORMS",         "1.3.5.CRS"),
    ("1.3.REPORTS & FORMS",         "1.3.5.Package Tracker"),
    ("1.3.REPORTS & FORMS",         "1.3.6.Monthly Report(As per QIC template)"),
    ("1.3.REPORTS & FORMS",         "1.3.7.Site Progress Model"),
    ("1.3.REPORTS & FORMS",         "1.3.8.4D (Frequency Upon approval)"),
    ("1.3.REPORTS & FORMS",         "1.3.9.MIDP"),
    ("1.3.REPORTS & FORMS",         "1.3.10.Laser Scan-ReCap Files"),
    ("1.3.REPORTS & FORMS",         "1.3.11.ACC Issues Dashboard"),
    ("1.3.REPORTS & FORMS",         "1.3.12_As Built Model"),
]

# ── Theme ─────────────────────────────────────────────────────────────────────
C = dict(
    bg="#F0F4FF", panel="#FFFFFF", panel2="#E8EEFF",
    accent="#1565C0", text="#0D1B2A", muted="#546E7A",
    border="#90CAF9", ok="#00897B", warn="#F57F17",
    err="#C62828", hover="#BBDEFB", rowa="#FFFFFF", rowb="#E3F2FD",
)
FM = ("Consolas", 9)
FS = ("Segoe UI", 8)

# ── Logic helpers ─────────────────────────────────────────────────────────────
def to_tree(rows):
    tree = {}
    for p, c in rows:
        p, c = p.strip(), c.strip()
        tree.setdefault(p, {})
        if c:
            tree[p][c] = {}
    return tree

def make_dirs(base, tree, res):
    for p, kids in tree.items():
        pp = base / p
        try:
            ex = pp.exists()
            pp.mkdir(parents=True, exist_ok=True)
            res["skipped" if ex else "created"].append(str(pp))
        except Exception as e:
            res["errors"].append(str(e))
        for k in kids:
            cp = pp / k
            try:
                ex = cp.exists()
                cp.mkdir(parents=True, exist_ok=True)
                res["skipped" if ex else "created"].append(str(cp))
            except Exception as e:
                res["errors"].append(str(e))

# ── CSV ───────────────────────────────────────────────────────────────────────
def csv_export(rows, path):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Parent Folder", "Sub Folder"])
        w.writerows(rows)

def csv_import(path):
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for i, r in enumerate(csv.reader(f)):
            if i == 0:
                continue
            if len(r) >= 2:
                rows.append((r[0].strip(), r[1].strip()))
            elif r and r[0].strip():
                rows.append((r[0].strip(), ""))
    return rows

# ── Excel ─────────────────────────────────────────────────────────────────────
def xlsx_export(rows, path):
    if not XLSX_OK:
        raise RuntimeError("pip install openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Folder Structure"
    th = Side(style="thin", color="CCCCCC")
    br = Border(left=th, right=th, top=th, bottom=th)
    hf = PatternFill("solid", fgColor="1F3864")
    hfont = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
    for col, h in enumerate(["Parent Folder", "Sub Folder"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont; c.fill = hf; c.border = br
        c.alignment = Alignment(horizontal="center", vertical="center")
    fills = [PatternFill("solid", fgColor="EBF1FF"),
             PatternFill("solid", fgColor="FFFFFF")]
    dfont = Font(name="Consolas", size=9)
    for i, (p, ch) in enumerate(rows, 2):
        for col, val in enumerate([p, ch], 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = dfont; c.fill = fills[i % 2]
            c.border = br; c.alignment = Alignment(vertical="center")
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 52
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2["A1"] = "HOW TO USE"
    ws2["A1"].font = Font(bold=True, size=12, name="Segoe UI")
    for r, n in enumerate([
        "", "Column A  →  Parent Folder  (top-level)",
        "Column B  →  Sub Folder     (child inside parent)", "",
        "• One row per sub-folder",
        "• Leave Column B empty if parent has no sub-folders",
        "• Do not edit the header row",
        "", "Import back:  File ▸ Import Excel",
    ], 2):
        ws2.cell(row=r, column=1, value=n).font = Font(name="Segoe UI", size=10)
    ws2.column_dimensions["A"].width = 55
    wb.save(path)

def xlsx_import(path):
    if not XLSX_OK:
        raise RuntimeError("pip install openpyxl")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = next((wb[s] for s in wb.sheetnames
               if "instruction" not in s.lower()), wb.active)
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        p  = str(row[0]).strip() if row[0] else ""
        ch = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if p:
            rows.append((p, ch))
    return rows

# ── App ───────────────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("BIM Folder Creator")
        self.configure(bg=C["bg"])
        self.geometry("860x700")
        self.minsize(680, 520)
        self.base_path = tk.StringVar()
        self.rows = list(DEFAULT_ROWS)
        self._build_menu()
        self._build_ui()
        self._refresh()
        self._center()

    def _center(self):
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        w, h   = self.winfo_width(), self.winfo_height()
        self.geometry("+{}+{}".format((sw - w) // 2, (sh - h) // 2))

    # ── Menu ──────────────────────────────────────────────────────────────────
    def _build_menu(self):
        def menu(parent, **kw):
            return tk.Menu(parent, tearoff=0, bg=C["panel"], fg=C["text"],
                           activebackground=C["accent"], activeforeground="white", **kw)
        mb = menu(self, relief="flat")
        self.config(menu=mb)

        fm = menu(mb)
        mb.add_cascade(label="File", menu=fm)
        fm.add_command(label="📥  Import CSV…",          command=self._imp_csv)
        fm.add_command(label="📥  Import Excel (.xlsx)…", command=self._imp_xlsx)
        fm.add_separator()
        fm.add_command(label="📤  Export CSV…",           command=self._exp_csv)
        fm.add_command(label="📤  Export Excel (.xlsx)…", command=self._exp_xlsx)
        fm.add_separator()
        fm.add_command(label="↺   Reset to Default",      command=self._reset)

        em = menu(mb)
        mb.add_cascade(label="Edit", menu=em)
        em.add_command(label="➕  Add Row",    command=self._add)
        em.add_command(label="🗑  Delete Row", command=self._delete)
        em.add_command(label="✕  Clear All",  command=self._clear)

    # ── UI ────────────────────────────────────────────────────────────────────
    def _build_ui(self):
        tk.Frame(self, bg=C["accent"], height=4).pack(fill="x")

        # Header
        hdr = tk.Frame(self, bg=C["bg"], pady=14)
        hdr.pack(fill="x", padx=24)
        tk.Label(hdr, text="BIM Folder Creator",
                 font=("Segoe UI", 17, "bold"),
                 bg=C["bg"], fg=C["text"]).pack(anchor="w")
        tk.Label(hdr, text="Edit table · Import / Export CSV or Excel · Create on disk",
                 font=FS, bg=C["bg"], fg=C["muted"]).pack(anchor="w", pady=(2, 0))
        tk.Label(hdr, text="Prepared by: Ahmed Khalaf",
                 font=("Segoe UI", 8, "italic"),
                 bg=C["bg"], fg=C["accent"]).pack(anchor="w", pady=(2, 0))

        tk.Frame(self, bg=C["border"], height=1).pack(fill="x", padx=24)

        # Toolbar
        tb = tk.Frame(self, bg=C["bg"], pady=10)
        tb.pack(fill="x", padx=24)

        def tbtn(text, cmd):
            tk.Button(tb, text=text, command=cmd, font=FS,
                      bg=C["panel"], fg=C["text"], relief="flat", bd=0,
                      padx=11, pady=6, cursor="hand2",
                      activebackground=C["hover"], activeforeground=C["text"]
                      ).pack(side="left", padx=(0, 5))

        tbtn("📥 Import CSV",    self._imp_csv)
        tbtn("📥 Import Excel",  self._imp_xlsx)
        tk.Frame(tb, bg=C["border"], width=1, height=20).pack(side="left", padx=5)
        tbtn("📤 Export CSV",    self._exp_csv)
        tbtn("📤 Export Excel",  self._exp_xlsx)
        tk.Frame(tb, bg=C["border"], width=1, height=20).pack(side="left", padx=5)
        tbtn("➕ Add Row",       self._add)
        tbtn("🗑 Delete Row",    self._delete)
        tbtn("↺ Reset",          self._reset)

        # Table header row
        lr = tk.Frame(self, bg=C["bg"])
        lr.pack(fill="x", padx=24, pady=(4, 3))
        tk.Label(lr, text="Folder Structure   —   double-click a cell to edit",
                 font=("Segoe UI", 9, "bold"),
                 bg=C["bg"], fg=C["muted"]).pack(side="left")
        self.count_lbl = tk.Label(lr, text="", font=FS,
                                   bg=C["bg"], fg=C["muted"])
        self.count_lbl.pack(side="right")

        # Treeview
        tbl = tk.Frame(self, bg=C["border"], bd=1)
        tbl.pack(fill="both", expand=True, padx=24)

        sty = ttk.Style(self)
        sty.theme_use("default")
        sty.configure("T.Treeview", background=C["panel"], foreground=C["text"],
                      fieldbackground=C["panel"], borderwidth=0,
                      font=FM, rowheight=24)
        sty.configure("T.Treeview.Heading", background=C["panel2"],
                      foreground=C["muted"],
                      font=("Segoe UI", 9, "bold"), relief="flat")
        sty.map("T.Treeview",
                background=[("selected", C["accent"])],
                foreground=[("selected", "white")])

        self.tv = ttk.Treeview(tbl, columns=("p", "c"), show="headings",
                               style="T.Treeview", selectmode="browse")
        self.tv.heading("p", text="Parent Folder")
        self.tv.heading("c", text="Sub Folder")
        self.tv.column("p", width=400, minwidth=180)
        self.tv.column("c", width=400, minwidth=180)
        self.tv.tag_configure("odd",  background=C["rowa"])
        self.tv.tag_configure("even", background=C["rowb"])
        self.tv.bind("<Double-1>", self._edit_cell)

        vsb = ttk.Scrollbar(tbl, orient="vertical",   command=self.tv.yview)
        hsb = ttk.Scrollbar(tbl, orient="horizontal", command=self.tv.xview)
        self.tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tv.pack(fill="both", expand=True)

        # Base directory row
        pr = tk.Frame(self, bg=C["bg"], pady=10)
        pr.pack(fill="x", padx=24)
        tk.Label(pr, text="Base Directory:", font=FS,
                 bg=C["bg"], fg=C["muted"]).pack(side="left", padx=(0, 8))
        tk.Entry(pr, textvariable=self.base_path, font=FM,
                 bg=C["panel"], fg=C["text"], insertbackground=C["text"],
                 relief="flat", bd=0, highlightthickness=1,
                 highlightbackground=C["border"],
                 highlightcolor=C["accent"]
                 ).pack(side="left", fill="x", expand=True, ipady=6, padx=(0, 8))
        tk.Button(pr, text="Browse…", font=FS,
                  bg=C["panel"], fg=C["text"], relief="flat", bd=0,
                  padx=10, pady=5, cursor="hand2",
                  activebackground=C["hover"], activeforeground=C["text"],
                  command=self._browse).pack(side="left")

        tk.Frame(self, bg=C["border"], height=1).pack(fill="x", padx=24)

        # Log + Run
        bot = tk.Frame(self, bg=C["bg"], pady=10)
        bot.pack(fill="x", padx=24)

        lf = tk.Frame(bot, bg=C["panel"], highlightthickness=1,
                      highlightbackground=C["border"])
        lf.pack(side="left", fill="both", expand=True, padx=(0, 12))
        self.log = tk.Text(lf, height=5, font=FM, bg=C["panel"], fg=C["text"],
                           relief="flat", bd=0, state="disabled", wrap="none")
        lsb = ttk.Scrollbar(lf, orient="vertical", command=self.log.yview)
        lsb.pack(side="right", fill="y")
        self.log.pack(fill="both", expand=True, padx=6, pady=4)
        self.log.tag_configure("ok",   foreground=C["ok"])
        self.log.tag_configure("skip", foreground=C["warn"])
        self.log.tag_configure("err",  foreground=C["err"])
        self.log.tag_configure("info", foreground=C["muted"])

        tk.Button(bot, text="▶  Create Folders",
                  font=("Segoe UI", 11, "bold"),
                  bg=C["accent"], fg="white", relief="flat", bd=0,
                  padx=22, pady=10, cursor="hand2",
                  activebackground="#0D47A1", activeforeground="white",
                  command=self._run).pack(side="right", anchor="s")

    # ── Table helpers ─────────────────────────────────────────────────────────
    def _refresh(self):
        self.tv.delete(*self.tv.get_children())
        for i, (p, c) in enumerate(self.rows):
            self.tv.insert("", "end", iid=str(i), values=(p, c),
                           tags=("odd" if i % 2 == 0 else "even",))
        self.count_lbl.config(text="{} rows · {} parents".format(
            len(self.rows), len({r[0] for r in self.rows})))

    def _edit_cell(self, event):
        if self.tv.identify_region(event.x, event.y) != "cell":
            return
        col = self.tv.identify_column(event.x)
        iid = self.tv.identify_row(event.y)
        if not iid:
            return
        ci  = 0 if col == "#1" else 1
        bb  = self.tv.bbox(iid, col)
        if not bb:
            return
        x, y, w, h = bb
        cur = self.tv.item(iid, "values")[ci]
        e = tk.Entry(self.tv, font=FM, bg=C["hover"], fg=C["text"],
                     insertbackground=C["text"], relief="flat",
                     highlightthickness=1, highlightbackground=C["accent"])
        e.place(x=x, y=y, width=w, height=h)
        e.insert(0, cur); e.focus_set(); e.select_range(0, "end")
        def commit(ev=None):
            vals = list(self.tv.item(iid, "values"))
            vals[ci] = e.get().strip()
            self.tv.item(iid, values=vals)
            self.rows[int(iid)] = (vals[0], vals[1])
            e.destroy()
        e.bind("<Return>", commit)
        e.bind("<FocusOut>", commit)
        e.bind("<Escape>", lambda ev: e.destroy())

    def _add(self):
        self.rows.append(("New Parent", "New Sub Folder"))
        self._refresh()
        last = str(len(self.rows) - 1)
        self.tv.selection_set(last); self.tv.see(last)

    def _delete(self):
        sel = self.tv.selection()
        if not sel:
            messagebox.showwarning("Select a row", "Click a row first.", parent=self)
            return
        del self.rows[int(sel[0])]
        self._refresh()

    def _clear(self):
        if messagebox.askyesno("Clear All", "Remove all rows?", parent=self):
            self.rows.clear(); self._refresh()

    def _reset(self):
        if messagebox.askyesno("Reset", "Reset to default structure?", parent=self):
            self.rows = list(DEFAULT_ROWS); self._refresh()

    # ── Import ────────────────────────────────────────────────────────────────
    def _imp_csv(self):
        p = filedialog.askopenfilename(title="Import CSV",
            filetypes=[("CSV", "*.csv"), ("All", "*.*")])
        if not p: return
        try:
            self.rows = csv_import(p); self._refresh()
            self._log("✓ Imported {} rows from {}".format(
                len(self.rows), Path(p).name), "ok")
        except Exception as ex:
            messagebox.showerror("Import Error", str(ex), parent=self)

    def _imp_xlsx(self):
        if not XLSX_OK:
            messagebox.showerror("Missing", "Run:  pip install openpyxl", parent=self); return
        p = filedialog.askopenfilename(title="Import Excel",
            filetypes=[("Excel", "*.xlsx"), ("All", "*.*")])
        if not p: return
        try:
            self.rows = xlsx_import(p); self._refresh()
            self._log("✓ Imported {} rows from {}".format(
                len(self.rows), Path(p).name), "ok")
        except Exception as ex:
            messagebox.showerror("Import Error", str(ex), parent=self)

    # ── Export ────────────────────────────────────────────────────────────────
    def _exp_csv(self):
        p = filedialog.asksaveasfilename(title="Export CSV",
            defaultextension=".csv", filetypes=[("CSV", "*.csv")])
        if not p: return
        try:
            csv_export(self.rows, p)
            self._log("✓ Saved → {}".format(p), "ok")
        except Exception as ex:
            messagebox.showerror("Export Error", str(ex), parent=self)

    def _exp_xlsx(self):
        if not XLSX_OK:
            messagebox.showerror("Missing", "Run:  pip install openpyxl", parent=self); return
        p = filedialog.asksaveasfilename(title="Export Excel",
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not p: return
        try:
            xlsx_export(self.rows, p)
            self._log("✓ Saved → {}".format(p), "ok")
            if os.name == "nt": os.startfile(p)
        except Exception as ex:
            messagebox.showerror("Export Error", str(ex), parent=self)

    # ── Browse & Run ──────────────────────────────────────────────────────────
    def _browse(self):
        p = filedialog.askdirectory(title="Select Base Directory")
        if p: self.base_path.set(p)

    def _log(self, msg, tag=""):
        self.log.configure(state="normal")
        self.log.insert("end", msg + "\n", tag)
        self.log.see("end")
        self.log.configure(state="disabled")

    def _run(self):
        base = self.base_path.get().strip()
        if not base:
            messagebox.showwarning("No Path", "Select a base directory first.", parent=self); return
        if not self.rows:
            messagebox.showwarning("Empty", "Folder list is empty.", parent=self); return
        if not Path(base).exists():
            messagebox.showerror("Not Found", "Path does not exist:\n{}".format(base), parent=self); return
        self.log.configure(state="normal")
        self.log.delete("1.0", "end")
        self.log.configure(state="disabled")
        self._log("[{}]  Creating…".format(datetime.now().strftime("%H:%M:%S")), "info")
        res = {"created": [], "skipped": [], "errors": []}
        make_dirs(Path(base), to_tree(self.rows), res)
        for p in res["created"]: self._log("✓  {}".format(p), "ok")
        for p in res["skipped"]: self._log("⚠  {}".format(p), "skip")
        for e in res["errors"]:  self._log("✗  {}".format(e), "err")
        c, s, e = len(res["created"]), len(res["skipped"]), len(res["errors"])
        self._log("\nDone — {} created  {} skipped  {} errors".format(c, s, e), "info")
        if (res["created"] or res["skipped"]) and os.name == "nt":
            os.startfile(base)

# ── Entry ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    App().mainloop()